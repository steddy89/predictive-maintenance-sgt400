import openpyxl
import sys

path = r'C:\Users\teddysudewo\OneDrive - Microsoft\TeddyS\My_git_project\applying-asynchronous-programming-c-sharp\temp_data.xlsx'
wb = openpyxl.load_workbook(path, read_only=True)
print('Sheets:', wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n--- {sheet_name} ---')
    rows = list(ws.iter_rows(max_row=6, values_only=True))
    for r in rows:
        print(r)
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
wb.close()
